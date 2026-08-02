"""
run_scan.py — المنسّق الرئيسي
يشغّل الماسح + يولّد Excel + يرسل تنبيه Telegram + يحفظ النتائج
"""
import os
import sys
import json
import io
import requests
from datetime import datetime

if sys.platform == 'win32':
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    else:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from full_market_whale_scanner import WhaleScanner, get_current_session
from self_learning import load_memory, save_memory, analyze_misses, daily_report


def generate_excel(signals, session_name):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[-] openpyxl غير مثبت — تخطي إنشاء Excel")
        return None

    wb = openpyxl.Workbook()

    # --- Sheet 1: كل الإشارات ---
    ws = wb.active
    ws.title = "الإشارات"
    headers = [
        "الرمز", "السعر", "Z-Score", "حجم نسبي", "قوة التجميع", "OBV",
        "RSI", "انكماش Bollinger", "شذوذ AI", "عدد الإشارات", "الإشارات"
    ]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row_idx, sig in enumerate(signals, 2):
        vd = sig.get('volume_data', {})
        acc = sig.get('accumulation', {})
        bb = sig.get('bollinger', {})
        sig_types = [s['type'] for s in sig.get('signals', [])]

        row_data = [
            sig['symbol'],
            round(sig.get('price', 0), 2),
            vd.get('z_score', 0),
            f"{vd.get('relative_volume', 0)}x",
            acc.get('cmf', 0),
            acc.get('obv_trend', ''),
            sig.get('rsi', 50),
            'نعم' if bb.get('squeeze') else 'لا',
            'نعم' if sig.get('is_anomaly') else 'لا',
            len(sig.get('signals', [])),
            ' + '.join(sig_types),
        ]

        # Color coding
        z = vd.get('z_score', 0)
        if z > 3:
            fill = PatternFill(start_color="ff6b6b", end_color="ff6b6b", fill_type="solid")
        elif z > 2:
            fill = PatternFill(start_color="ffa94d", end_color="ffa94d", fill_type="solid")
        else:
            fill = PatternFill(start_color="69db7c", end_color="69db7c", fill_type="solid")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if col in [3, 4, 5]:
                cell.fill = fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # --- Sheet 2: خيارات غير عادية ---
    ws2 = wb.create_sheet("خيارات غير عادية")
    opt_headers = ["الرمز", "العقد", "النوع", "السعر", "الصالة", "الحجم", "Open Interest", "النسبة"]
    for col, h in enumerate(opt_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    opt_row = 2
    for sig in signals:
        for signal in sig.get('signals', []):
            if signal['type'] == 'UNUSUAL_OPTIONS':
                opt_data = signal.get('options_data', {})
                for contract in opt_data.get('contracts', []):
                    ws2.cell(row=opt_row, column=1, value=sig['symbol']).border = border
                    ws2.cell(row=opt_row, column=2, value=contract.get('contract', '')).border = border
                    ws2.cell(row=opt_row, column=3, value=contract.get('type', '')).border = border
                    ws2.cell(row=opt_row, column=4, value=contract.get('strike', 0)).border = border
                    ws2.cell(row=opt_row, column=5, value=contract.get('expiry', '')).border = border
                    ws2.cell(row=opt_row, column=6, value=contract.get('volume', 0)).border = border
                    ws2.cell(row=opt_row, column=7, value=contract.get('open_interest', 0)).border = border
                    ws2.cell(row=opt_row, column=8, value=f"{contract.get('ratio', 0)}x").border = border
                    opt_row += 1

    for col in range(1, len(opt_headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # --- Sheet 3: بيع عَمَي ---
    ws3 = wb.create_sheet("بيع عَمَي")
    short_headers = ["الرمز", "السعر", "نسبة بيع العَمَي", "أيام التغطية", "العوامة"]
    for col, h in enumerate(short_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    short_row = 2
    for sig in signals:
        for signal in sig.get('signals', []):
            if signal['type'] == 'HIGH_SHORT_INTEREST':
                sd = signal.get('short_data', {})
                ws3.cell(row=short_row, column=1, value=sig['symbol']).border = border
                ws3.cell(row=short_row, column=2, value=round(sig.get('price', 0), 2)).border = border
                ws3.cell(row=short_row, column=3, value=f"{sd.get('short_percent', 0)*100:.1f}%").border = border
                ws3.cell(row=short_row, column=4, value=sd.get('days_to_cover', 0)).border = border
                ws3.cell(row=short_row, column=5, value=f"{sd.get('float_shares', 0)/1e6:.1f}M").border = border
                short_row += 1

    for col in range(1, len(short_headers) + 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Save
    filename = f"whale_scan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"[+] Excel: {filename}")
    return filename


def send_telegram(signals, session_name):
    token = os.environ.get('TELEGRAM_BOT_TOKEN', '')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID', '')
    if not token or not chat_id:
        print("[-] Telegram غير مُعدّ — تخطي")
        return

    if not signals:
        msg = f"🐋 ماسح الحيتان\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n⏱ {session_name}\n\nلا توجد إشارات مثيرة اليوم."
    else:
        msg = f"🐋 ماسح الحيتان\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n⏱ {session_name}\n\n📊 {len(signals)} سهم بإشارات:\n\n"

        for i, sig in enumerate(signals[:10], 1):
            vd = sig.get('volume_data', {})
            acc = sig.get('accumulation', {})
            sig_types = [s['type'] for s in sig.get('signals', [])]
            icon = '🔴' if len(sig_types) >= 3 else '🟡' if len(sig_types) >= 2 else '🟢'
            msg += f"{icon} {i}. {sig['symbol']} — ${sig.get('price', 0):.2f}\n"
            msg += f"   Z={vd.get('z_score', 0)} | حجم={vd.get('relative_volume', 0)}x | قوة التجميع={acc.get('cmf', 0)}\n"
            msg += f"   {' + '.join(sig_types)}\n\n"

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = requests.post(url, json={
            "chat_id": chat_id,
            "text": msg,
            "parse_mode": "HTML",
        }, timeout=10)
        if resp.status_code == 200:
            print(f"[+] Telegram: تم الإرسال")
        else:
            print(f"[-] Telegram خطأ: {resp.status_code}")
    except Exception as e:
        print(f"[-] Telegram خطأ: {e}")


def main():
    scanner = WhaleScanner()
    signals = scanner.scan()

    session_code, session_name = get_current_session()

    output = {
        'scan_time': datetime.now().isoformat(),
        'session': session_code,
        'session_name': session_name,
        'total_signals': len(signals),
        'signals': signals,
    }
    with open('scan_results.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)
    print(f"[+] محفوظ في scan_results.json")

    # Excel
    generate_excel(signals, session_name)

    # Telegram
    send_telegram(signals, session_name)

    # Self-learning analysis
    memory = load_memory()
    hit, missed = analyze_misses(memory)
    save_memory(memory)

    if missed:
        print(f"\n[!] فُقد {len(missed)} أسهم صعدت:{', '.join(m['symbol'] for m in missed[:5])}")
        report = daily_report(memory)
        print(report)

    print(f"\n[✓] اكتمل المسح — {len(signals)} إشارة")


if __name__ == "__main__":
    main()
